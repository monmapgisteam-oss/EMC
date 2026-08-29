# -*- coding: utf-8 -*-
"""
«mes1-12 нэгтгэл 2023 ArcGIS.xlsx» -> data/excel.js

sheet «Нийт орд arcgis» бүтэц:
  Сар | Түвшин | Блок | Үзүүлэлт | ... 15 утгын багана (Excel D..R)

  Түлхүүр бүр (Сар, Түвшин, Блок) яг 5 мөртэй:
    Нийт хэмжээ (мян.т) · Cu агуулга (%) · Cu металл (т)
    Mo агуулга (%)      · Mo металл (т)

Гаралт: public/data/excel.json = { "Сар|Түвшин|Блок": {kt:[...15], cu:[...], cut:[...], mo:[...], mot:[...]} }

Ажиллуулах:  python tools/build_excel.py
"""
import os, io, json, collections
import openpyxl

HERE  = os.path.dirname(os.path.abspath(__file__))
ROOT  = os.path.dirname(HERE)
XLSX  = os.environ.get("EMC_XLSX", r"I:\EMC\mes1-12 negtgel 2023 ArcGIS.xlsx")
SHEET = "Нийт орд arcgis"
OUT   = os.path.join(ROOT, "public", "data", "excel.json")

MEAS = {
    "Нийт хэмжээ (мян.т)": "kt",
    "Cu агуулга (%)":      "cu",
    "Cu металл (т)":       "cut",
    "Mo агуулга (%)":      "mo",
    "Mo металл (т)":       "mot",
}
DEC = {"kt": 2, "cu": 4, "cut": 2, "mo": 5, "mot": 3}
FIRST_COL, LAST_COL = 4, 19          # 0-индекс: D..R


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit("«%s» sheet олдсонгүй. Байгаа sheet-үүд: %s" % (SHEET, wb.sheetnames))
    ws = wb[SHEET]

    rows = [r[:LAST_COL] for r in ws.iter_rows(max_col=LAST_COL, values_only=True)]
    body = [r for r in rows[1:] if r[0] is not None]

    groups = collections.OrderedDict()
    unknown = set()
    for r in body:
        label = str(r[3])
        if label not in MEAS:
            unknown.add(label)
            continue
        key = "%d|%d|%s" % (int(r[0]), int(r[1]), str(r[2]))
        which = MEAS[label]
        groups.setdefault(key, {})[which] = [
            round(r[i], DEC[which]) if isinstance(r[i], (int, float)) and abs(r[i]) > 1e-9 else 0
            for i in range(FIRST_COL, LAST_COL)
        ]

    if unknown:
        print("АНХААР — танигдаагүй «Үзүүлэлт»:", sorted(unknown))

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    text = json.dumps(groups, ensure_ascii=False, separators=(",", ":"))
    with io.open(OUT, "w", encoding="utf-8") as f:
        f.write(text)

    months = sorted({int(k.split("|")[0]) for k in groups})
    levels = sorted({int(k.split("|")[1]) for k in groups}, reverse=True)
    blocks = sorted({k.split("|")[2] for k in groups})
    print("бичив public/data/excel.json  (%d бүлэг, %.1f KB)" % (len(groups), len(text.encode("utf-8")) / 1024))
    print("  сар   :", months)
    print("  түвшин:", levels)
    print("  блок  :", blocks)
    wb.close()


if __name__ == "__main__":
    main()
